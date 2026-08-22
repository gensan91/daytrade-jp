"""
tools/export_backtest.py  — 「バックテスト」シートを data/backtest.csv に書き出す
使い方: py tools/export_backtest.py [xlsmファイル]
デフォルト: py tools/export_backtest.py デイトレ用.xlsm

実行タイミング：
  BacktestGap200 (VBA) 実行・保存後にこのスクリプトを実行する。
  出力: data/backtest.csv (BOM付きUTF-8 / CRLF)

出力内容（L列）:
  証券コード / 銘柄名 / 日付 / 前日終値 / 始値 / 高値 / 安値 / 終値
  / ギャップ率(%) / TP目標 / SL目標 / 日足仮想勝敗

サマリーも末尾に出力する。
"""
import sys, os, csv
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')


def fmt_cell(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime('%Y-%m-%d')
        return v.strftime('%H:%M:%S')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    return str(v)


def main():
    xlsm = sys.argv[1] if len(sys.argv) > 1 else 'デイトレ用.xlsm'

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl が必要です: py -m pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(xlsm):
        print(f"ERROR: {xlsm} が見つかりません")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsm, data_only=True)

    if 'バックテスト' not in wb.sheetnames:
        print("ERROR: 「バックテスト」シートが存在しません。")
        print("  先に BacktestGap200 (VBA) を実行してから再試行してください。")
        sys.exit(1)

    ws = wb['バックテスト']
    os.makedirs('data', exist_ok=True)
    out_path = 'data/backtest.csv'

    headers = [cell.value for cell in ws[1]]
    if not any(headers):
        print("ERROR: バックテストシートのヘッダが空です。VBAの実行結果を確認してください。")
        sys.exit(1)

    # データ行（空行またはサマリー行まで）
    data_rows = []
    summary_rows = []
    in_summary = False

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 空行を挟んだ後の行はサマリー
        if all(v is None for v in row):
            continue
        first_val = str(row[0]) if row[0] is not None else ''
        if 'サマリー' in first_val or in_summary:
            in_summary = True
            summary_rows.append(row)
            continue
        # 日付が入っていればデータ行
        if row[2] is not None:
            data_rows.append(row)

    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, lineterminator='\r\n')
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow([fmt_cell(v) for v in row])

    print(f"data/backtest.csv を生成: {len(data_rows)} 件のギャップ日")

    # サマリー集計（CSV からカウント）
    verdicts = [row[11] for row in data_rows if row[11] is not None]
    cnt_win  = verdicts.count('利確')
    cnt_loss = verdicts.count('損切')
    cnt_und  = verdicts.count('未達')
    cnt_unk  = verdicts.count('不明')
    total_det = cnt_win + cnt_loss

    print("\n=== BacktestGap200 サマリー ===")
    print(f"  ギャップ3%以上の日数: {len(data_rows)}")
    print(f"  利確（確定）:         {cnt_win}")
    print(f"  損切（確定）:         {cnt_loss}")
    print(f"  未達:                 {cnt_und}")
    print(f"  不明（順序依存）:     {cnt_unk}")
    if total_det > 0:
        win_rate = cnt_win / total_det * 100
        print(f"  利確率（確定分）:     {win_rate:.1f}%  (n={total_det})")

    # 銘柄別集計
    from collections import defaultdict
    by_code = defaultdict(lambda: {'win': 0, 'loss': 0, 'und': 0, 'unk': 0})
    for row in data_rows:
        code = str(row[0]) if row[0] else ''
        v = str(row[11]) if row[11] else ''
        if v == '利確':   by_code[code]['win']  += 1
        elif v == '損切': by_code[code]['loss'] += 1
        elif v == '未達': by_code[code]['und']  += 1
        else:             by_code[code]['unk']  += 1

    print("\n=== 銘柄別（ギャップ日あり）===")
    for code in sorted(by_code.keys()):
        d = by_code[code]
        det = d['win'] + d['loss']
        wr = f"{d['win']/det*100:.0f}%" if det > 0 else '-'
        total_g = d['win'] + d['loss'] + d['und'] + d['unk']
        print(f"  {code}: 利確={d['win']} 損切={d['loss']} 未達={d['und']} 不明={d['unk']}"
              f"  /  確定分勝率={wr}  (n={total_g})")


if __name__ == '__main__':
    main()
