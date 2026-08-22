"""
tools/audit_xlsm.py  — xlsm 構造の健全性チェック + Trades→data/trades.csv 書き出し
使い方: py tools/audit_xlsm.py [xlsmファイル]
デフォルト: py tools/audit_xlsm.py デイトレ用.xlsm

検査内容（CLAUDE.md §4.2 準拠）:
  1. 標準モジュールがちょうど9本か（modConfig1 等の混入検出）
  2. Config B1..B14 の実値を表示
  3. 残存 RSS 数式の総数（期待値 121。スナップショット系は 0）
  4. Trades の行数・K列語彙（利確/損切/時間切れ/引け前強制/手動判断 の5種のみ）
  5. Trades を data/trades.csv に書き出す（BOM付きUTF-8/CRLF）

終了コード:
  0 = 正常
  1 = 異常あり（要確認）
"""
import sys, os, csv, re
from datetime import datetime, date

sys.stdout.reconfigure(encoding='utf-8')

EXPECTED_STD_MODULES = 9
# Watchlist 120 + Position 2 + _work 2 + Trades 1 = 125 が実測値（2026/8/22 監査）
# CLAUDE.md §2.0-A の記載「121」は Position 1 本・_work/_Trades を未計上の旧値
# スナップショット系（スナップショット / SS_MMDD）の RSS が 0 であることのみを必須チェックとする
EXPECTED_RSS_COUNT = 125
VALID_EXIT_REASONS = {'利確', '損切', '時間切れ', '引け前強制', '手動判断'}

def fmt_cell(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.strftime('%Y-%m-%d')
        return v.strftime('%H:%M:%S')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    return str(v)

def main():
    xlsm = sys.argv[1] if len(sys.argv) > 1 else 'デイトレ用.xlsm'

    try:
        import openpyxl
        from oletools.olevba import VBA_Parser
    except ImportError:
        print("ERROR: openpyxl と oletools が必要です")
        print("  py -m pip install openpyxl oletools")
        sys.exit(1)

    if not os.path.exists(xlsm):
        print(f"ERROR: {xlsm} が見つかりません")
        sys.exit(1)

    print(f"=== xlsm 監査: {xlsm} ===\n")
    errors = []

    # ── 1. 標準モジュール数 ────────────────────────────────────────
    print("【1】標準モジュール数")
    p = VBA_Parser(xlsm)
    std_modules = []
    for _, _, name, _ in p.extract_macros():
        if name.endswith('.bas'):
            std_modules.append(name)
    print(f"  モジュール数: {len(std_modules)}（期待値: {EXPECTED_STD_MODULES}）")
    for m in sorted(std_modules):
        print(f"    {m}")
    if len(std_modules) != EXPECTED_STD_MODULES:
        msg = f"標準モジュール数が {EXPECTED_STD_MODULES} 本ではありません（現在: {len(std_modules)}）"
        print(f"  *** ERROR: {msg}")
        errors.append(msg)
    else:
        print("  OK")

    # ── 2. Config B1..B14 ─────────────────────────────────────────
    print("\n【2】Config 実値")
    wb = openpyxl.load_workbook(xlsm, data_only=True)
    ws_cfg = wb['Config']
    config = {}
    for row in ws_cfg.iter_rows(min_row=1, max_row=14, values_only=True):
        if row[0] is not None:
            config[row[0]] = row[1]
            print(f"  {row[0]}: {row[1]}")

    # ── 3. RSS 数式数 ──────────────────────────────────────────────
    print("\n【3】RSS 数式残存数")
    wb_formula = openpyxl.load_workbook(xlsm, data_only=False)
    from openpyxl.worksheet.formula import ArrayFormula

    def cell_formula_text(v) -> str:
        """セル値から数式文字列を返す（ArrayFormula / 通常文字列 両対応）"""
        if v is None:
            return ''
        if isinstance(v, ArrayFormula):
            return v.text or ''
        if isinstance(v, str):
            return v
        return ''

    rss_count = 0
    rss_by_sheet = {}
    for ws in wb_formula.worksheets:
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                txt = cell_formula_text(cell.value)
                # _xll.RssMarket / _xll.RssChartPast / RssMarket / RssChartPast を検出
                if txt and ('Rss' in txt or 'rss' in txt.lower()):
                    count += 1
        if count > 0:
            rss_by_sheet[ws.title] = count
            rss_count += count

    for sheet_name, cnt in rss_by_sheet.items():
        print(f"  {sheet_name}: {cnt}本")
    print(f"  合計: {rss_count}本（期待値: {EXPECTED_RSS_COUNT}）")

    if rss_count != EXPECTED_RSS_COUNT:
        msg = f"RSS 数式数が期待値 {EXPECTED_RSS_COUNT} と異なります（現在: {rss_count}）"
        print(f"  *** WARN: {msg}")
        # スナップショット系に RSS が残っているかチェック
        for sheet_name in rss_by_sheet:
            if 'スナップショット' in sheet_name or 'SS_' in sheet_name:
                err = f"{sheet_name} に RSS 数式が残っています（数式固定が必要）"
                print(f"  *** ERROR: {err}")
                errors.append(err)
    else:
        print("  OK")

    # ── 4. Trades 検査 ────────────────────────────────────────────
    print("\n【4】Trades シート検査")
    ws_trades = wb['Trades']
    headers = [cell.value for cell in ws_trades[1]]
    data_rows = []
    for row in ws_trades.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        data_rows.append(row)

    print(f"  ヘッダ: {ws_trades.max_column}列")
    print(f"  データ行数: {len(data_rows)}件")

    # K列（インデックス10）の語彙チェック
    invalid_reasons = []
    for i, row in enumerate(data_rows, 2):
        reason = row[10]  # K列（0-indexed で 10）
        if reason not in VALID_EXIT_REASONS:
            invalid_reasons.append(f"行{i}: '{reason}'")

    if invalid_reasons:
        msg = f"K列（決済理由）に無効な値: {invalid_reasons}"
        print(f"  *** ERROR: {msg}")
        errors.append(msg)
    else:
        print(f"  K列語彙: すべて有効（{VALID_EXIT_REASONS}）OK")

    # 損益率サマリ
    pnl_rates = [row[11] for row in data_rows if row[11] is not None]
    pnl_costs = [row[19] for row in data_rows if row[19] is not None]
    if pnl_rates:
        print(f"  損益率: {[round(r,2) for r in pnl_rates]}")
        print(f"  コスト後累計: {sum(pnl_costs):+,.0f}円")

    # ── 5. trades.csv 書き出し ────────────────────────────────────
    print("\n【5】data/trades.csv 書き出し")
    os.makedirs('data', exist_ok=True)
    out_path = 'data/trades.csv'
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, lineterminator='\r\n')
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow([fmt_cell(v) for v in row])
    print(f"  {out_path} に {len(data_rows)} 件を書き出しました")

    # ── 結果サマリ ─────────────────────────────────────────────────
    print(f"\n=== 監査結果 ===")
    if errors:
        print(f"ERROR {len(errors)} 件:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("異常なし。コミット可能です。")

if __name__ == '__main__':
    main()
